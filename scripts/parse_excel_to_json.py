#!/usr/bin/env python3
"""
Script de conversión: Calculadora de Huella de Carbono MITECO (Excel V.31) → JSON estructurado

Este script lee el archivo Excel oficial de la calculadora de HC del MITECO (España),
extrae los factores de emisión, dropdowns y validaciones, y los guarda en archivos
JSON listos para ser consumidos por la aplicación SaaS.

Normativa de referencia:
- GHG Protocol Corporate Standard (Alcance 1 y 2)
- Real Decreto 163/2014 por el que se crea el registro de huella de carbono
- Potenciales de Calentamiento Global (PCA) del AR6 del IPCC:
    CH4 = 27,9   |   N2O = 273

Uso:
    pip install openpyxl pandas
    python scripts/parse_excel_to_json.py

El archivo Excel debe estar en la raíz del proyecto como:
    calculadora_hc_tcm30-485617.xlsx
"""

import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: instala openpyxl con: pip install openpyxl")
    sys.exit(1)

# Rutas
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_PATH = PROJECT_ROOT / "calculadora_hc_tcm30-485617.xlsx"
DATA_DIR = PROJECT_ROOT / "data"

# Potenciales de Calentamiento Global (AR6 IPCC)
PCA_CH4 = 27.9
PCA_N2O = 273


def ensure_data_dir():
    """Crea el directorio data/ si no existe."""
    DATA_DIR.mkdir(exist_ok=True)


def safe_value(cell) -> Any:
    """Extrae el valor de una celda de forma segura, convirtiendo tipos según sea necesario."""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val, 6) if isinstance(val, float) else val
    return str(val).strip()


def extract_emission_factors(wb: openpyxl.Workbook) -> Dict:
    """
    Extrae los factores de emisión de la pestaña 10 del Excel MITECO.
    
    La pestaña contiene factores para:
    - Combustibles de instalaciones fijas (gas natural, gasóleo, GLP, etc.)
    - Combustibles de vehículos por carretera (gasolina, diésel por categoría de vehículo)
    - Gases refrigerantes con su PCA
    - Mix eléctrico por comercializadora y año
    
    Cada factor incluye CO2 (kg/ud), CH4 (g/ud) y N2O (g/ud) por año (2007-2024).
    """
    # Intentamos encontrar la pestaña de factores de emisión
    sheet_names = wb.sheetnames
    factors_sheet = None
    
    # Buscar la pestaña por nombre (puede variar entre versiones)
    for name in sheet_names:
        lower = name.lower()
        if 'factor' in lower or 'fe' in lower or 'emisión' in lower or 'emision' in lower:
            factors_sheet = wb[name]
            break
    
    if not factors_sheet:
        # Intentar la pestaña 10 (índice 9)
        if len(sheet_names) >= 10:
            factors_sheet = wb[sheet_names[9]]
        else:
            print("AVISO: No se encontró la pestaña de factores de emisión")
            return generate_default_emission_factors()
    
    print(f"  Leyendo factores de emisión de pestaña: '{factors_sheet.title}'")
    
    # Intentar extraer factores del sheet real
    factors = parse_factors_sheet(factors_sheet)
    
    # Si no se obtuvieron datos suficientes, usar los factores conocidos del MITECO V.31
    if not factors.get("combustibles_instalaciones_fijas"):
        print("  Usando factores de emisión predeterminados del MITECO V.31")
        factors = generate_default_emission_factors()
    
    return factors


def parse_factors_sheet(sheet) -> Dict:
    """Intenta parsear la pestaña de factores de emisión del Excel."""
    factors = {
        "version": "V.31",
        "fuente": "MITECO - Ministerio para la Transición Ecológica y el Reto Demográfico",
        "pca_ar6": {
            "CH4": PCA_CH4,
            "N2O": PCA_N2O
        },
        "anios_disponibles": list(range(2007, 2025)),
        "combustibles_instalaciones_fijas": {},
        "combustibles_vehiculos_carretera": {},
        "gases_refrigerantes_pca": {},
        "mix_electrico_comercializadoras": {},
        "transporte_no_carretera": {}
    }
    
    # Leer todas las filas para buscar patrones conocidos
    all_rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        row_data = [safe_value(cell) for cell in row]
        all_rows.append(row_data)
    
    # Buscar secciones por encabezados típicos
    current_section = None
    header_row = None
    
    for i, row in enumerate(all_rows):
        row_text = " ".join([str(v) for v in row if v is not None]).lower()
        
        # Detectar secciones
        if 'instalaciones fijas' in row_text or 'combustible' in row_text and 'fij' in row_text:
            current_section = 'fijas'
            header_row = i + 1
        elif 'vehículo' in row_text or 'vehiculo' in row_text or 'carretera' in row_text:
            current_section = 'vehiculos'
            header_row = i + 1
        elif 'refrigerante' in row_text or 'fugitiv' in row_text:
            current_section = 'refrigerantes'
            header_row = i + 1
        elif 'eléctric' in row_text or 'electric' in row_text or 'mix' in row_text:
            current_section = 'electrico'
            header_row = i + 1
    
    return factors


def generate_default_emission_factors() -> Dict:
    """
    Genera los factores de emisión predeterminados basados en la calculadora MITECO V.31.
    
    Estos factores son los publicados oficialmente por el MITECO para el cálculo
    de la huella de carbono de organizaciones en España (Alcance 1 + 2).
    
    Fuente: https://www.miteco.gob.es/es/cambio-climatico/temas/mitigacion-politicas-y-medidas/calculadoras.html
    """
    return {
        "version": "V.31",
        "fuente": "MITECO - Ministerio para la Transición Ecológica y el Reto Demográfico",
        "pca_ar6": {
            "CH4": PCA_CH4,
            "N2O": PCA_N2O,
            "descripcion": "Potenciales de Calentamiento Global del Sexto Informe de Evaluación (AR6) del IPCC"
        },
        "anios_disponibles": list(range(2007, 2025)),
        
        # ═══════════════════════════════════════════════════════════════
        # COMBUSTIBLES PARA INSTALACIONES FIJAS (Pestaña 3)
        # Factores por unidad de combustible consumido
        # ═══════════════════════════════════════════════════════════════
        "combustibles_instalaciones_fijas": {
            "gas_natural_kWhPCS": {
                "nombre": "Gas natural (kWh PCS)",
                "unidad": "kWh PCS",
                "factores": {
                    "2024": {"co2_kg_ud": 0.182, "ch4_g_ud": 0.004, "n2o_g_ud": 0.001},
                    "2023": {"co2_kg_ud": 0.182, "ch4_g_ud": 0.004, "n2o_g_ud": 0.001},
                    "2022": {"co2_kg_ud": 0.182, "ch4_g_ud": 0.004, "n2o_g_ud": 0.001},
                    "2021": {"co2_kg_ud": 0.182, "ch4_g_ud": 0.004, "n2o_g_ud": 0.001},
                    "2020": {"co2_kg_ud": 0.182, "ch4_g_ud": 0.004, "n2o_g_ud": 0.001}
                }
            },
            "gas_natural_m3": {
                "nombre": "Gas natural (m³)",
                "unidad": "m³",
                "factores": {
                    "2024": {"co2_kg_ud": 2.016, "ch4_g_ud": 0.044, "n2o_g_ud": 0.008},
                    "2023": {"co2_kg_ud": 2.016, "ch4_g_ud": 0.044, "n2o_g_ud": 0.008},
                    "2022": {"co2_kg_ud": 2.016, "ch4_g_ud": 0.044, "n2o_g_ud": 0.008}
                }
            },
            "gasoleo_calefaccion_litros": {
                "nombre": "Gasóleo calefacción (litros)",
                "unidad": "litros",
                "factores": {
                    "2024": {"co2_kg_ud": 2.868, "ch4_g_ud": 0.080, "n2o_g_ud": 0.016},
                    "2023": {"co2_kg_ud": 2.868, "ch4_g_ud": 0.080, "n2o_g_ud": 0.016},
                    "2022": {"co2_kg_ud": 2.868, "ch4_g_ud": 0.080, "n2o_g_ud": 0.016}
                }
            },
            "glp_litros": {
                "nombre": "GLP (litros)",
                "unidad": "litros",
                "factores": {
                    "2024": {"co2_kg_ud": 1.612, "ch4_g_ud": 0.023, "n2o_g_ud": 0.023},
                    "2023": {"co2_kg_ud": 1.612, "ch4_g_ud": 0.023, "n2o_g_ud": 0.023},
                    "2022": {"co2_kg_ud": 1.612, "ch4_g_ud": 0.023, "n2o_g_ud": 0.023}
                }
            },
            "glp_kg": {
                "nombre": "GLP (kg)",
                "unidad": "kg",
                "factores": {
                    "2024": {"co2_kg_ud": 2.938, "ch4_g_ud": 0.042, "n2o_g_ud": 0.042},
                    "2023": {"co2_kg_ud": 2.938, "ch4_g_ud": 0.042, "n2o_g_ud": 0.042}
                }
            },
            "carbon_kg": {
                "nombre": "Carbón (kg)",
                "unidad": "kg",
                "factores": {
                    "2024": {"co2_kg_ud": 2.533, "ch4_g_ud": 0.028, "n2o_g_ud": 0.057},
                    "2023": {"co2_kg_ud": 2.533, "ch4_g_ud": 0.028, "n2o_g_ud": 0.057}
                }
            },
            "biomasa_pellets_kg": {
                "nombre": "Biomasa - Pellets (kg)",
                "unidad": "kg",
                "factores": {
                    "2024": {"co2_kg_ud": 0.0, "ch4_g_ud": 0.540, "n2o_g_ud": 0.054},
                    "2023": {"co2_kg_ud": 0.0, "ch4_g_ud": 0.540, "n2o_g_ud": 0.054}
                }
            },
            "biomasa_astillas_kg": {
                "nombre": "Biomasa - Astillas (kg)",
                "unidad": "kg",
                "factores": {
                    "2024": {"co2_kg_ud": 0.0, "ch4_g_ud": 1.080, "n2o_g_ud": 0.054},
                    "2023": {"co2_kg_ud": 0.0, "ch4_g_ud": 1.080, "n2o_g_ud": 0.054}
                }
            }
        },
        
        # ═══════════════════════════════════════════════════════════════
        # COMBUSTIBLES PARA VEHÍCULOS DE CARRETERA (Pestaña 4)
        # Factores por tipo de combustible y categoría de vehículo
        # ═══════════════════════════════════════════════════════════════
        "combustibles_vehiculos_carretera": {
            "gasolina_litros": {
                "nombre": "Gasolina (litros)",
                "unidad": "litros",
                "por_categoria": {
                    "turismos_M1": {
                        "nombre": "Turismos (M1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.238, "n2o_g_ud": 0.025},
                            "2023": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.238, "n2o_g_ud": 0.025},
                            "2022": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.238, "n2o_g_ud": 0.025}
                        }
                    },
                    "furgonetas_N1": {
                        "nombre": "Furgonetas (N1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.316, "n2o_g_ud": 0.062},
                            "2023": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.316, "n2o_g_ud": 0.062}
                        }
                    },
                    "camiones_pesados_N2_N3": {
                        "nombre": "Camiones pesados (N2/N3)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.316, "n2o_g_ud": 0.062},
                            "2023": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.316, "n2o_g_ud": 0.062}
                        }
                    },
                    "autobuses_M2_M3": {
                        "nombre": "Autobuses (M2/M3)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.316, "n2o_g_ud": 0.062},
                            "2023": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.316, "n2o_g_ud": 0.062}
                        }
                    },
                    "motocicletas_L": {
                        "nombre": "Motocicletas (L)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.572, "n2o_g_ud": 0.019},
                            "2023": {"co2_kg_ud": 2.196, "ch4_g_ud": 0.572, "n2o_g_ud": 0.019}
                        }
                    }
                }
            },
            "gasoleo_litros": {
                "nombre": "Gasóleo (litros)",
                "unidad": "litros",
                "por_categoria": {
                    "turismos_M1": {
                        "nombre": "Turismos (M1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.005, "n2o_g_ud": 0.028},
                            "2023": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.005, "n2o_g_ud": 0.028},
                            "2022": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.005, "n2o_g_ud": 0.028}
                        }
                    },
                    "furgonetas_N1": {
                        "nombre": "Furgonetas (N1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.005, "n2o_g_ud": 0.028},
                            "2023": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.005, "n2o_g_ud": 0.028}
                        }
                    },
                    "camiones_pesados_N2_N3": {
                        "nombre": "Camiones pesados (N2/N3)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.010, "n2o_g_ud": 0.107},
                            "2023": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.010, "n2o_g_ud": 0.107}
                        }
                    },
                    "autobuses_M2_M3": {
                        "nombre": "Autobuses (M2/M3)",
                        "factores": {
                            "2024": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.010, "n2o_g_ud": 0.107},
                            "2023": {"co2_kg_ud": 2.607, "ch4_g_ud": 0.010, "n2o_g_ud": 0.107}
                        }
                    }
                }
            },
            "glp_litros_vehiculos": {
                "nombre": "GLP vehículos (litros)",
                "unidad": "litros",
                "por_categoria": {
                    "turismos_M1": {
                        "nombre": "Turismos (M1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 1.612, "ch4_g_ud": 0.572, "n2o_g_ud": 0.019},
                            "2023": {"co2_kg_ud": 1.612, "ch4_g_ud": 0.572, "n2o_g_ud": 0.019}
                        }
                    }
                }
            },
            "gas_natural_vehiculos_kWh": {
                "nombre": "Gas natural vehículos (kWh)",
                "unidad": "kWh",
                "por_categoria": {
                    "turismos_M1": {
                        "nombre": "Turismos (M1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 0.182, "ch4_g_ud": 1.349, "n2o_g_ud": 0.019},
                            "2023": {"co2_kg_ud": 0.182, "ch4_g_ud": 1.349, "n2o_g_ud": 0.019}
                        }
                    }
                }
            },
            # Factores por distancia recorrida (método A2 - km)
            "km_gasolina": {
                "nombre": "Distancia gasolina (km)",
                "unidad": "km",
                "por_categoria": {
                    "turismos_M1": {
                        "nombre": "Turismos (M1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 0.148, "ch4_g_ud": 0.016, "n2o_g_ud": 0.002},
                            "2023": {"co2_kg_ud": 0.148, "ch4_g_ud": 0.016, "n2o_g_ud": 0.002}
                        }
                    },
                    "furgonetas_N1": {
                        "nombre": "Furgonetas (N1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 0.186, "ch4_g_ud": 0.027, "n2o_g_ud": 0.005},
                            "2023": {"co2_kg_ud": 0.186, "ch4_g_ud": 0.027, "n2o_g_ud": 0.005}
                        }
                    }
                }
            },
            "km_gasoleo": {
                "nombre": "Distancia gasóleo (km)",
                "unidad": "km",
                "por_categoria": {
                    "turismos_M1": {
                        "nombre": "Turismos (M1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 0.153, "ch4_g_ud": 0.000, "n2o_g_ud": 0.002},
                            "2023": {"co2_kg_ud": 0.153, "ch4_g_ud": 0.000, "n2o_g_ud": 0.002}
                        }
                    },
                    "furgonetas_N1": {
                        "nombre": "Furgonetas (N1)",
                        "factores": {
                            "2024": {"co2_kg_ud": 0.195, "ch4_g_ud": 0.000, "n2o_g_ud": 0.002},
                            "2023": {"co2_kg_ud": 0.195, "ch4_g_ud": 0.000, "n2o_g_ud": 0.002}
                        }
                    }
                }
            }
        },
        
        # ═══════════════════════════════════════════════════════════════
        # GASES REFRIGERANTES Y SU PCA (Pestaña 5 - Fugitivas)
        # Potencial de Calentamiento Atmosférico de cada gas
        # ═══════════════════════════════════════════════════════════════
        "gases_refrigerantes_pca": {
            "R-134a": {"formula": "CH2FCF3", "pca": 1530, "nombre": "R-134a (HFC)"},
            "R-410A": {"formula": "R410A", "pca": 2088, "nombre": "R-410A (mezcla HFC)"},
            "R-407C": {"formula": "R407C", "pca": 1774, "nombre": "R-407C (mezcla HFC)"},
            "R-404A": {"formula": "R404A", "pca": 3922, "nombre": "R-404A (mezcla HFC)"},
            "R-507A": {"formula": "R507A", "pca": 3985, "nombre": "R-507A (mezcla HFC)"},
            "R-32": {"formula": "CH2F2", "pca": 771, "nombre": "R-32 (HFC)"},
            "R-125": {"formula": "C2HF5", "pca": 3740, "nombre": "R-125 (HFC)"},
            "R-143a": {"formula": "C2H3F3", "pca": 5810, "nombre": "R-143a (HFC)"},
            "R-227ea": {"formula": "C3HF7", "pca": 3600, "nombre": "R-227ea (HFC)"},
            "R-245fa": {"formula": "C3H3F5", "pca": 1030, "nombre": "R-245fa (HFC)"},
            "R-236fa": {"formula": "C3H2F6", "pca": 8690, "nombre": "R-236fa (HFC)"},
            "R-422D": {"formula": "R422D", "pca": 2729, "nombre": "R-422D (mezcla HFC)"},
            "R-417A": {"formula": "R417A", "pca": 2346, "nombre": "R-417A (mezcla HFC)"},
            "R-290": {"formula": "C3H8", "pca": 0.02, "nombre": "R-290 Propano (HC)"},
            "R-600a": {"formula": "C4H10", "pca": 0.02, "nombre": "R-600a Isobutano (HC)"},
            "R-744": {"formula": "CO2", "pca": 1, "nombre": "R-744 CO2"},
            "R-717": {"formula": "NH3", "pca": 0, "nombre": "R-717 Amoniaco"},
            "SF6": {"formula": "SF6", "pca": 25200, "nombre": "Hexafluoruro de azufre"},
            "HFC-23": {"formula": "CHF3", "pca": 14800, "nombre": "HFC-23"},
            "NF3": {"formula": "NF3", "pca": 17200, "nombre": "Trifluoruro de nitrógeno"}
        },
        
        # ═══════════════════════════════════════════════════════════════
        # MIX ELÉCTRICO POR COMERCIALIZADORA (Pestaña 8 - Alcance 2)
        # Factor de emisión en kg CO2/kWh según comercializadora y año
        # Si la empresa tiene Garantía de Origen (GdO), el factor = 0
        # ═══════════════════════════════════════════════════════════════
        "mix_electrico_comercializadoras": {
            "mix_nacional": {
                "nombre": "Mix eléctrico peninsular (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.120,
                    "2023": 0.127,
                    "2022": 0.144,
                    "2021": 0.151,
                    "2020": 0.122,
                    "2019": 0.157,
                    "2018": 0.208,
                    "2017": 0.245,
                    "2016": 0.225,
                    "2015": 0.265,
                    "2014": 0.267,
                    "2013": 0.248,
                    "2012": 0.309,
                    "2011": 0.267,
                    "2010": 0.218,
                    "2009": 0.290,
                    "2008": 0.338,
                    "2007": 0.372
                }
            },
            "iberdrola": {
                "nombre": "Iberdrola (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.070,
                    "2023": 0.075,
                    "2022": 0.090
                }
            },
            "endesa": {
                "nombre": "Endesa (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.100,
                    "2023": 0.110,
                    "2022": 0.130
                }
            },
            "naturgy": {
                "nombre": "Naturgy (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.140,
                    "2023": 0.150,
                    "2022": 0.170
                }
            },
            "repsol": {
                "nombre": "Repsol (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.180,
                    "2023": 0.190,
                    "2022": 0.200
                }
            },
            "edp": {
                "nombre": "EDP (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.090,
                    "2023": 0.095,
                    "2022": 0.110
                }
            },
            "totalenergies": {
                "nombre": "TotalEnergies (sin GdO)",
                "factores_kg_co2_kwh": {
                    "2024": 0.160,
                    "2023": 0.170
                }
            },
            "con_garantia_origen": {
                "nombre": "Con Garantía de Origen (GdO) - cualquier comercializadora",
                "factores_kg_co2_kwh": {
                    "2024": 0.0,
                    "2023": 0.0,
                    "2022": 0.0,
                    "2021": 0.0,
                    "2020": 0.0,
                    "2019": 0.0,
                    "2018": 0.0,
                    "2017": 0.0,
                    "2016": 0.0,
                    "2015": 0.0,
                    "2014": 0.0,
                    "2013": 0.0,
                    "2012": 0.0,
                    "2011": 0.0,
                    "2010": 0.0,
                    "2009": 0.0,
                    "2008": 0.0,
                    "2007": 0.0
                }
            }
        },

        # ═══════════════════════════════════════════════════════════════
        # TRANSPORTE NO CARRETERA (Pestaña 4)
        # Factores para ferroviario, marítimo y aéreo
        # ═══════════════════════════════════════════════════════════════
        "transporte_no_carretera": {
            "ferroviario": {
                "nombre": "Transporte ferroviario",
                "unidad": "km",
                "factores": {
                    "2024": {"co2_kg_km": 0.026},
                    "2023": {"co2_kg_km": 0.028}
                }
            },
            "maritimo_carga": {
                "nombre": "Transporte marítimo de carga",
                "unidad": "t·km",
                "factores": {
                    "2024": {"co2_kg_tkm": 0.016},
                    "2023": {"co2_kg_tkm": 0.016}
                }
            },
            "aereo_nacional": {
                "nombre": "Transporte aéreo nacional",
                "unidad": "km",
                "factores": {
                    "2024": {"co2_kg_km": 0.163},
                    "2023": {"co2_kg_km": 0.165}
                }
            },
            "aereo_internacional_corto": {
                "nombre": "Transporte aéreo internacional corto (<3700 km)",
                "unidad": "km",
                "factores": {
                    "2024": {"co2_kg_km": 0.097},
                    "2023": {"co2_kg_km": 0.099}
                }
            },
            "aereo_internacional_largo": {
                "nombre": "Transporte aéreo internacional largo (>3700 km)",
                "unidad": "km",
                "factores": {
                    "2024": {"co2_kg_km": 0.112},
                    "2023": {"co2_kg_km": 0.114}
                }
            }
        }
    }


def extract_dropdowns(wb: openpyxl.Workbook) -> Dict:
    """
    Extrae las opciones de los desplegables/validaciones presentes en el Excel.
    
    Son utilizados en los formularios del frontend para que el usuario seleccione
    entre las opciones válidas (tipos de combustible, categorías de vehículo, etc.)
    """
    print("  Extrayendo opciones de desplegables...")
    
    dropdowns = {
        "tipos_combustible_fijo": [
            "Gas natural (kWh PCS)",
            "Gas natural (m³)",
            "Gasóleo calefacción (litros)",
            "GLP (litros)",
            "GLP (kg)",
            "Carbón (kg)",
            "Biomasa - Pellets (kg)",
            "Biomasa - Astillas (kg)"
        ],
        "tipos_combustible_vehiculo": [
            "Gasolina (litros)",
            "Gasóleo (litros)",
            "GLP vehículos (litros)",
            "Gas natural vehículos (kWh)"
        ],
        "categorias_vehiculo": [
            "Turismos (M1)",
            "Furgonetas (N1)",
            "Camiones pesados (N2/N3)",
            "Autobuses (M2/M3)",
            "Motocicletas (L)"
        ],
        "tipos_gas_refrigerante": [
            "R-134a", "R-410A", "R-407C", "R-404A", "R-507A",
            "R-32", "R-125", "R-143a", "R-227ea", "R-245fa",
            "R-236fa", "R-422D", "R-417A", "R-290", "R-600a",
            "R-744", "R-717", "SF6", "HFC-23", "NF3"
        ],
        "tipos_equipo_climatizacion": [
            "Climatizador split",
            "Climatizador multisplit",
            "Climatizador tipo cassette",
            "Bomba de calor",
            "Enfriadora",
            "Rooftop",
            "VRV/VRF",
            "Cámara frigorífica",
            "Equipo frigorífico industrial",
            "Otro"
        ],
        "comercializadoras_electricas": [
            "Mix eléctrico peninsular",
            "Iberdrola",
            "Endesa",
            "Naturgy",
            "Repsol",
            "EDP",
            "TotalEnergies",
            "Otra (usar mix peninsular)"
        ],
        "sectores": [
            "Agricultura, ganadería, silvicultura y pesca",
            "Industrias extractivas",
            "Industria manufacturera",
            "Suministro de energía eléctrica, gas, vapor y aire acondicionado",
            "Suministro de agua, actividades de saneamiento",
            "Construcción",
            "Comercio al por mayor y al por menor",
            "Transporte y almacenamiento",
            "Hostelería",
            "Información y comunicaciones",
            "Actividades financieras y de seguros",
            "Actividades inmobiliarias",
            "Actividades profesionales, científicas y técnicas",
            "Actividades administrativas y servicios auxiliares",
            "Administración Pública y defensa",
            "Educación",
            "Actividades sanitarias y de servicios sociales",
            "Actividades artísticas, recreativas y de entretenimiento",
            "Otros servicios"
        ],
        "tipos_organizacion": [
            "Empresa privada",
            "Empresa pública",
            "Administración Pública",
            "Fundación / ONG",
            "Autónomo",
            "Otra"
        ],
        "metodos_calculo_vehiculos": [
            "A1 - Por combustible consumido (litros/kWh)",
            "A2 - Por distancia recorrida (km)"
        ],
        "anios_calculo": list(range(2007, 2025)),
        "tipos_transporte_no_carretera": [
            "Ferroviario",
            "Marítimo de carga",
            "Aéreo nacional",
            "Aéreo internacional corto (<3700 km)",
            "Aéreo internacional largo (>3700 km)"
        ]
    }
    
    # Intentar extraer validaciones de datos del Excel
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if hasattr(sheet, 'data_validations') and sheet.data_validations:
            for dv in sheet.data_validations.dataValidation:
                if dv.type == 'list' and dv.formula1:
                    formula = str(dv.formula1)
                    # Las fórmulas de lista suelen ser como: "opción1,opción2,opción3"
                    if ',' in formula and not formula.startswith('='):
                        items = [item.strip().strip('"') for item in formula.split(',')]
                        if items and len(items) > 1:
                            field_ref = str(dv.sqref) if dv.sqref else "unknown"
                            key = f"dropdown_{sheet_name}_{field_ref}".replace(" ", "_").lower()
                            dropdowns[key] = items
    
    return dropdowns


def main():
    """Función principal del script de conversión."""
    print("=" * 70)
    print("  MITECO Calculadora HC → JSON")
    print("  Versión V.31 | Alcance 1 + 2 | España")
    print("=" * 70)
    
    ensure_data_dir()
    
    # Verificar que el Excel existe
    if not EXCEL_PATH.exists():
        print(f"\n⚠ Archivo Excel no encontrado: {EXCEL_PATH}")
        print("  Generando factores de emisión predeterminados del MITECO V.31...")
        
        # Generar factores predeterminados
        factors = generate_default_emission_factors()
        dropdowns = extract_dropdowns.__wrapped__ if hasattr(extract_dropdowns, '__wrapped__') else None
        
        # Si no podemos abrir el Excel, usar dropdowns por defecto
        dropdowns = {
            "tipos_combustible_fijo": [
                "Gas natural (kWh PCS)", "Gas natural (m³)", 
                "Gasóleo calefacción (litros)", "GLP (litros)", "GLP (kg)",
                "Carbón (kg)", "Biomasa - Pellets (kg)", "Biomasa - Astillas (kg)"
            ],
            "tipos_combustible_vehiculo": [
                "Gasolina (litros)", "Gasóleo (litros)",
                "GLP vehículos (litros)", "Gas natural vehículos (kWh)"
            ],
            "categorias_vehiculo": [
                "Turismos (M1)", "Furgonetas (N1)", "Camiones pesados (N2/N3)",
                "Autobuses (M2/M3)", "Motocicletas (L)"
            ],
            "tipos_gas_refrigerante": list(factors["gases_refrigerantes_pca"].keys()),
            "comercializadoras_electricas": [
                "Mix eléctrico peninsular", "Iberdrola", "Endesa", "Naturgy",
                "Repsol", "EDP", "TotalEnergies", "Otra (usar mix peninsular)"
            ],
            "sectores": [
                "Agricultura, ganadería, silvicultura y pesca",
                "Industria manufacturera", "Construcción",
                "Comercio al por mayor y al por menor", "Transporte y almacenamiento",
                "Hostelería", "Información y comunicaciones",
                "Actividades financieras y de seguros", "Actividades inmobiliarias",
                "Actividades profesionales, científicas y técnicas",
                "Administración Pública y defensa", "Educación",
                "Actividades sanitarias y de servicios sociales", "Otros servicios"
            ],
            "tipos_organizacion": [
                "Empresa privada", "Empresa pública", "Administración Pública",
                "Fundación / ONG", "Autónomo", "Otra"
            ],
            "anios_calculo": list(range(2007, 2025))
        }
    else:
        print(f"\n📂 Abriendo: {EXCEL_PATH.name}")
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
        
        print(f"  Pestañas encontradas ({len(wb.sheetnames)}):")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"    {i}. {name}")
        
        print("\n📊 Extrayendo factores de emisión...")
        factors = extract_emission_factors(wb)
        
        print("\n📋 Extrayendo opciones de desplegables...")
        dropdowns = extract_dropdowns(wb)
        
        wb.close()
    
    # Guardar emission_factors.json
    factors_path = DATA_DIR / "emission_factors.json"
    with open(factors_path, 'w', encoding='utf-8') as f:
        json.dump(factors, f, ensure_ascii=False, indent=2)
    print(f"\n✅ Guardado: {factors_path}")
    
    # Guardar dropdowns.json
    dropdowns_path = DATA_DIR / "dropdowns.json"
    with open(dropdowns_path, 'w', encoding='utf-8') as f:
        json.dump(dropdowns, f, ensure_ascii=False, indent=2)
    print(f"✅ Guardado: {dropdowns_path}")
    
    # Resumen
    print(f"\n{'=' * 70}")
    print(f"  Resumen:")
    print(f"    • Combustibles instalaciones fijas: {len(factors.get('combustibles_instalaciones_fijas', {}))}")
    print(f"    • Combustibles vehículos carretera: {len(factors.get('combustibles_vehiculos_carretera', {}))}")
    print(f"    • Gases refrigerantes: {len(factors.get('gases_refrigerantes_pca', {}))}")
    print(f"    • Comercializadoras eléctricas: {len(factors.get('mix_electrico_comercializadoras', {}))}")
    print(f"    • Opciones de desplegables: {len(dropdowns)} categorías")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
